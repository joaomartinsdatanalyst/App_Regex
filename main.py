import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import regex

# Função para a Tela 01
def tela01():
    def next_step():
        #next_step() lê o texto da área de texto, divide-o em linhas e armazena em uma variável global dados.
        global dados
        #4 - Obtém o texto da área de texto, remove espaços em branco no início/fim e divide por linhas
        dados = text_area.get("1.0", tk.END).strip().split('\n') #transforma em uma lista
        #5 - Fecha a janela atual
        root.destroy()
        #6 - Chama a função para a Tela 02
        tela02()
    
    # 1 - Uma janela é criada usando Tkinter
    root = tk.Tk()
    root.title("Tela 01: Entrada de Dados")
    
    # Cria e exibe um rótulo (label)
    label = tk.Label(root, text="Cole as células do Excel aqui:")
    label.pack(pady=10)
    
    #2 - Cria e exibe uma área de texto (INPUT DOS DADOS)
    text_area = tk.Text(root, width=50, height=20)
    text_area.pack(pady=10)
    
    #3 - Cria e exibe um botão que chama next_step ao ser clicado
    #Quando o usuário clica no botão "Próximo", a função next_step() é chamada.
    next_button = tk.Button(root, text="Próximo", command=next_step)
    next_button.pack(pady=10)
    
    # Inicia o loop principal da janela
    # root.mainloop() inicia o loop principal do evento, que é um ciclo contínuo que aguarda e responde a eventos (como cliques de botões, digitação de texto, etc.).
    # Sem este loop, a janela do Tkinter seria exibida e fechada imediatamente, pois o programa terminaria sua execução.
    # Mantém a Janela Ativa:
    root.mainloop()

# Função para a Tela 02
def tela02():
    def processar(): #trabalhar aqui
        #processar() lê os padrões definidos pelo usuário e aplica esses padrões aos dados coletados na Tela 01.
        # Obtém os valores dos campos de entrada e dos comboboxes
        padrao1 = entry1.get() #1 padrão do usuario
        padrao2 = entry2.get() #2 padrão do usuario
        tipo1 = combo1.get() #1 tipo de padrão
        tipo2 = combo2.get() #2 tipo de padrão
        
        #TO DO: fazer uma funcao para cada um dos tres cenarios
        #TO DO: fazer condicionais para ativar cada funcao
        if padrao1 == '' and padrao2 == '':
            nova_execucao('erro')


        n_cel = len(dados) #quantos dados foram inseridos na primeira tela
        for i in range(0, n_cel):
            if padrao1 in dados[i] and padrao2 == '':
                resultado = regex.comeca_com_e_nao_tem_x_caracteres(dados[i],padrao1)
                resultado = ";".join(resultado)
                continue
            elif padrao2 in dados[i] and padrao1 == '':
                resultado = regex.comeca_com_e_nao_tem_x_caracteres(dados[i],padrao2)
                resultado = ";".join(resultado)
                continue
            elif padrao1 in dados[i] and int(padrao2) > 0:
                resultado = regex.comeca_com_e_tem_x_caracteres(dados[i],padrao1,int(padrao2))
                resultado = ";".join(resultado)
                continue
            elif padrao2 in dados[i] and int(padrao1) > 0:
                resultado = regex.comeca_com_e_tem_x_caracteres(dados[i],padrao2,int(padrao1))
                resultado = ";".join(resultado)
                continue
            else:
                continue
            
        # resultados = []
        # for linha in dados:
        #     resultado = ''
        #     # Verifica se a linha corresponde ao padrão 1
        #     if (tipo1 == "Começa com" and linha.startswith(padrao1)) or (tipo1 == "Tem X caracteres" and len(linha) == int(padrao1)):
        #         resultado = linha
        #     # Verifica se a linha corresponde ao padrão 2, se fornecido
        #     if padrao2 and ((tipo2 == "Começa com" and linha.startswith(padrao2)) or (tipo2 == "Tem X caracteres" and len(linha) == int(padrao2))):
        #         resultado = linha
        #     # Adiciona o resultado à lista de resultados
        #     resultados.append(resultado)
        
        # Chama a função para exportar os resultados para o Excel
        exportar_excel(dados, 'resultados', root)
    
    #7 - Cria uma nova janela Tkinter
    root = tk.Tk()
    root.title("Tela 02: Definição de Padrões")
    
    #8 - Cria e exibe os campos de entrada e comboboxes para o Padrão 1
    label1 = tk.Label(root, text="Padrão 1:")
    label1.grid(row=0, column=0, padx=10, pady=5)
    entry1 = tk.Entry(root)
    entry1.grid(row=0, column=1, padx=10, pady=5)
    
    combo1 = ttk.Combobox(root, values=["Começa com", "Tem X caracteres"])
    combo1.grid(row=0, column=2, padx=10, pady=5)
    combo1.current(0)
    
    #9 - Cria e exibe os campos de entrada e comboboxes para o Padrão 2 (opcional)
    label2 = tk.Label(root, text="Padrão 2 (Opcional):")
    label2.grid(row=1, column=0, padx=10, pady=5)
    entry2 = tk.Entry(root)
    entry2.grid(row=1, column=1, padx=10, pady=5)
    
    combo2 = ttk.Combobox(root, values=["Começa com", "Tem X caracteres"])
    combo2.grid(row=1, column=2, padx=10, pady=5)
    combo2.current(0)
    
    #10 - Cria e exibe um botão que chama processar ao ser clicado
    process_button = tk.Button(root, text="Processar", command=processar)
    process_button.grid(row=2, columnspan=3, pady=20)
    
    # Inicia o loop principal da janela
    root.mainloop()

# Função para exportar os dados para o Excel
def exportar_excel(dados, resultados):
    # Cria uma nova planilha de trabalho
    wb = openpyxl.Workbook()
    ws = wb.active
    # Preenche as colunas A e B com os dados e resultados
    for i, (dado, resultado) in enumerate(zip(dados, resultados), start=1):
        ws[f'A{i}'] = dado
        ws[f'B{i}'] = resultado
    # Salva a planilha como 'resultado.xlsx'
    wb.save('resultado.xlsx')
    # Exibe uma mensagem informando que a exportação foi concluída
    messagebox.showinfo("Exportação Concluída", "Os dados foram exportados para 'resultado.xlsx'.\nPrograma desenvolvido por [Seu Nome].")
    # Chama a função para oferecer a opção de nova execução
    nova_execucao('sucesso')

# Função para reiniciar ou fechar o programa
def nova_execucao(status, root):
    if status == 'erro':
        resposta = messagebox.askyesno("Ocorreu um erro, deseja tentar novamente?")
    else:
        # Pergunta ao usuário se deseja realizar outro processo
        resposta = messagebox.askyesno("Continuar?", "Deseja realizar outro processo?")
    if resposta:
        # Se sim, chama a Tela 01
        root.destroy()
        tela01()
    else:
        # Se não, encerra o programa
        exit()

# Inicializando o programa
if __name__ == "__main__":
    tela01()
